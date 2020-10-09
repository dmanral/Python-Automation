# Reading an excel file.

import openpyxl

# Opening excel workbook.
# Workbook object.
workbook = openpyxl.load_workbook('xml_files/example.xlsx')
print(type(workbook))

# If you already know the name of the individual sheet.
sheet = workbook.get_sheet_by_name('Sheet1')
print(type(sheet))

# If you dont know the name of the sheet. This will print all the sheets by name.
# print(workbook.get_sheet_names()) # Old way
print(workbook.sheetnames)          # New way

# Get shell object.
cell = sheet['A1']
# Get shell value.
print(cell.value)
print(str(cell.value))
print(str(sheet['A1'].value))

print(sheet['B1'].value)
print(sheet['C1'].value)        # You will get int because its int in cell.
print(str(sheet['C1'].value))   # You can convert int to string.

# By row and column, same as sheet['B1'] (object).
print(sheet.cell(row=1, column=2))

# Get values of a whole row.
for i in range(1, 8):
    # Will print B1-B8 values.
    print(i, sheet.cell(row=i, column=2).value)
