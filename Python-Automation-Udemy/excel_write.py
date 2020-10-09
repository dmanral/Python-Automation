# Writing an excel file.

import openpyxl
# Workbook object created.
wb = openpyxl.Workbook()

print(wb)

# Should automatically have one sheet.
print(wb.sheetnames)

# Sheet object
sheet = wb.get_sheet_by_name('Sheet')
print(sheet)

print(sheet['A1'])  # Should print none as it empty.

# Assign to cell.
sheet['A1'] = 42
sheet['A2'] = 'Hello'

# Saving new workbook.
wb.save('xml_files/example2_original.xlsx')

# Adding new worksheet to existing excel file.
sheet2 = wb.create_sheet()
print(wb.get_sheet_names())
print(sheet2.title)

# New name for sheet.
sheet2.title = 'My New Sheet Name'
print(wb.get_sheet_names())

# Saved modified into new excel.
wb.save('xml_files/example2_b.xlsx')

# Adding sheet to very beginning (depends on index).
wb.create_sheet(index=0, title='My Other Sheet')
wb.save('xml_files/example2_c.xlsx')
